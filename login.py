import streamlit as st
import openpyxl
import os

EXCEL_DB = "users.xlsx"

# ------------------ INITIALIZE EXCEL FILE ------------------
def init_excel():
    """Create users.xlsx with header if it does not exist."""
    if not os.path.exists(EXCEL_DB):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["username", "password"])  # header
        wb.save(EXCEL_DB)

# Call initializer
init_excel()

# ------------------ LOAD USERS ------------------
def load_users():
    """Load all users from the Excel file."""
    wb = openpyxl.load_workbook(EXCEL_DB)
    ws = wb.active

    users = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            users[row[0]] = {"pass": row[1]}
    return users

# ------------------ SAVE A NEW USER ------------------
def save_user(username, password):
    """Save a new user to Excel safely."""
    try:
        wb = openpyxl.load_workbook(EXCEL_DB)
        ws = wb.active
        ws.append([username, password])
        wb.save(EXCEL_DB)
    except PermissionError:
        st.error("❌ Close 'users.xlsx' file before adding new user.")

# ------------------ LOGIN PAGE UI ------------------
def login_page():
    st.title("🔐 Login to MindAI")

    users = load_users()

    tab1, tab2 = st.tabs(["Login", "Sign Up"])

    # ---------------- LOGIN ----------------
    with tab1:
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")

        if st.button("Login"):
            if username in users and users[username]["pass"] == password:
                st.success("Login Successful!")
                st.session_state.logged_in = True
                st.rerun()
            else:
                st.error("❌ Incorrect username or password")

    # ---------------- SIGN UP ----------------
    with tab2:
        new_user = st.text_input("New Username")
        new_pass = st.text_input("New Password", type="password")

        if st.button("Sign Up"):
            if new_user in users:
                st.warning("⚠ User already exists")
            else:
                save_user(new_user, new_pass)
                st.success("Account Created Successfully!")
